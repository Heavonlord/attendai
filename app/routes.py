from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from flask_login import login_user, logout_user, login_required, current_user
from app import db, bcrypt, socketio
from app.models import User, Course, Enrollment, Attendance
from datetime import datetime, date, timedelta
from functools import wraps
import csv
import io
import json

main = Blueprint('main', __name__)


# ─── Decorators ────────────────────────────────────────────────
def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('main.login'))
            if current_user.role not in roles:
                flash('Access denied.', 'danger')
                return redirect(url_for('main.dashboard'))
            return f(*args, **kwargs)
        return decorated
    return decorator


# ─── Auth Routes ───────────────────────────────────────────────
@main.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    return render_template('index.html')


@main.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        remember = request.form.get('remember') == 'on'

        user = User.query.filter(
            (User.username == username) | (User.email == username)
        ).first()

        if not user:
            flash('Invalid username or password.', 'danger')
            return render_template('login.html')

        # Check account lockout
        if user.locked_until and user.locked_until > datetime.utcnow():
            flash(f'Account locked. Try again after {user.locked_until.strftime("%H:%M")}.', 'danger')
            return render_template('login.html')

        if user.check_password(password) and user.is_active:
            user.failed_logins = 0
            user.locked_until = None
            db.session.commit()
            login_user(user, remember=remember)
            next_page = request.args.get('next')
            return redirect(next_page or url_for('main.dashboard'))
        else:
            user.failed_logins = (user.failed_logins or 0) + 1
            if user.failed_logins >= 5:
                user.locked_until = datetime.utcnow() + timedelta(minutes=30)
                flash('Too many failed attempts. Account locked for 30 minutes.', 'danger')
            else:
                flash(f'Invalid password. {5 - user.failed_logins} attempts remaining.', 'danger')
            db.session.commit()

    return render_template('login.html')


@main.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('main.login'))


# ─── Dashboard Router ──────────────────────────────────────────
@main.route('/dashboard')
@login_required
def dashboard():
    if current_user.role == 'admin':
        return redirect(url_for('main.admin_dashboard'))
    elif current_user.role == 'teacher':
        return redirect(url_for('main.teacher_dashboard'))
    elif current_user.role == 'parent':
        return redirect(url_for('main.parent_dashboard'))
    else:
        return redirect(url_for('main.student_dashboard'))


# ─── Admin Routes ──────────────────────────────────────────────
@main.route('/admin')
@login_required
@role_required('admin')
def admin_dashboard():
    users = User.query.all()
    courses = Course.query.all()
    students = [u for u in users if u.role == 'student']
    teachers = [u for u in users if u.role == 'teacher']

    stats = {
        'total_students': len(students),
        'total_teachers': len(teachers),
        'total_courses': len(courses),
        'total_attendance': Attendance.query.count()
    }

    return render_template('admin_dashboard.html', users=users, courses=courses,
                           stats=stats, students=students, teachers=teachers)


@main.route('/admin/users/add', methods=['POST'])
@login_required
@role_required('admin')
def add_user():
    data = request.form
    if User.query.filter_by(username=data['username']).first():
        flash('Username already exists.', 'danger')
        return redirect(url_for('main.admin_dashboard'))
    if User.query.filter_by(email=data['email']).first():
        flash('Email already exists.', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    user = User(
        username=data['username'],
        email=data['email'],
        role=data['role'],
        roll_no=data.get('roll_no') or None,
        parent_phone=data.get('parent_phone') or None,
        parent_email=data.get('parent_email') or None
    )
    user.set_password(data['password'])
    db.session.add(user)
    db.session.commit()
    flash(f'User {user.username} created successfully.', 'success')
    return redirect(url_for('main.admin_dashboard'))


@main.route('/admin/users/delete/<int:user_id>', methods=['POST'])
@login_required
@role_required('admin')
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash('Cannot delete yourself.', 'danger')
        return redirect(url_for('main.admin_dashboard'))
    db.session.delete(user)
    db.session.commit()
    flash(f'User {user.username} deleted.', 'success')
    return redirect(url_for('main.admin_dashboard'))


@main.route('/admin/users/bulk-upload', methods=['POST'])
@login_required
@role_required('admin')
def bulk_upload_users():
    file = request.files.get('csv_file')
    if not file:
        flash('No file uploaded.', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    stream = io.StringIO(file.stream.read().decode('UTF8'))
    reader = csv.DictReader(stream)
    added, failed = 0, []

    for i, row in enumerate(reader, 2):
        try:
            if User.query.filter_by(username=row['username']).first():
                failed.append(f"Row {i}: Username '{row['username']}' exists")
                continue
            if User.query.filter_by(email=row['email']).first():
                failed.append(f"Row {i}: Email '{row['email']}' exists")
                continue
            user = User(
                username=row['username'].strip(),
                email=row['email'].strip(),
                role=row.get('role', 'student').strip(),
                roll_no=row.get('roll_no', '').strip() or None,
                parent_phone=row.get('parent_phone', '').strip() or None,
                parent_email=row.get('parent_email', '').strip() or None
            )
            user.set_password(row.get('password', 'Pass@1234'))
            db.session.add(user)
            added += 1
        except Exception as e:
            failed.append(f"Row {i}: {str(e)}")

    db.session.commit()
    flash(f'Bulk upload: {added} users added, {len(failed)} failed.', 'success' if not failed else 'warning')
    if failed:
        session['bulk_errors'] = failed[:10]
    return redirect(url_for('main.admin_dashboard'))


@main.route('/admin/courses/add', methods=['POST'])
@login_required
@role_required('admin')
def add_course():
    data = request.form
    if Course.query.filter_by(code=data['code']).first():
        flash('Course code already exists.', 'danger')
        return redirect(url_for('main.admin_dashboard'))
    course = Course(
        name=data['name'],
        code=data['code'].upper(),
        teacher_id=int(data['teacher_id'])
    )
    db.session.add(course)
    db.session.commit()
    flash(f'Course {course.code} created.', 'success')
    return redirect(url_for('main.admin_dashboard'))


@main.route('/admin/courses/delete/<int:course_id>', methods=['POST'])
@login_required
@role_required('admin')
def delete_course(course_id):
    course = Course.query.get_or_404(course_id)
    db.session.delete(course)
    db.session.commit()
    flash(f'Course {course.code} deleted.', 'success')
    return redirect(url_for('main.admin_dashboard'))


@main.route('/admin/analytics')
@login_required
@role_required('admin')
def admin_analytics():
    students = User.query.filter_by(role='student').all()
    courses = Course.query.all()
    teachers = User.query.filter_by(role='teacher').all()

    risk_counts = {'SAFE': 0, 'CAUTION': 0, 'WARNING': 0, 'CRITICAL': 0}
    student_stats = []
    for s in students:
        pct = s.get_attendance_percentage()
        risk, _ = s.get_risk_level()
        risk_counts[risk] += 1
        student_stats.append({'student': s, 'pct': pct, 'risk': risk})

    overall_avg = sum(x['pct'] for x in student_stats) / len(student_stats) if student_stats else 0

    course_stats = []
    for c in courses:
        enrolled = len(list(c.enrollments))
        avg = c.get_avg_attendance()
        at_risk = sum(1 for s in c.get_enrolled_students()
                      if s.get_attendance_percentage(c.id) < 75)
        course_stats.append({'course': c, 'enrolled': enrolled, 'avg': avg, 'at_risk': at_risk})

    return render_template('admin_analytics.html',
                           risk_counts=risk_counts, student_stats=student_stats,
                           overall_avg=overall_avg, course_stats=course_stats,
                           teachers=teachers)


@main.route('/admin/csv-template')
@login_required
@role_required('admin')
def csv_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['username', 'email', 'password', 'role', 'roll_no', 'parent_phone', 'parent_email'])
    writer.writerow(['john_doe', 'john@example.com', 'Pass@1234', 'student', 'CS101', '+919876543210', 'parent@example.com'])
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode()), mimetype='text/csv',
                     as_attachment=True, download_name='users_template.csv')


# ─── Teacher Routes ────────────────────────────────────────────
@main.route('/teacher')
@login_required
@role_required('teacher')
def teacher_dashboard():
    courses = Course.query.filter_by(teacher_id=current_user.id).all()
    course_data = []
    for c in courses:
        enrolled = len(list(c.enrollments))
        avg = c.get_avg_attendance()
        total_classes = c.get_total_classes()
        course_data.append({'course': c, 'enrolled': enrolled, 'avg': avg, 'total_classes': total_classes})
    return render_template('teacher_dashboard.html', course_data=course_data)


@main.route('/teacher/course/<int:course_id>/enrollments', methods=['GET', 'POST'])
@login_required
@role_required('teacher')
def manage_enrollments(course_id):
    course = Course.query.get_or_404(course_id)
    if course.teacher_id != current_user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('main.teacher_dashboard'))

    if request.method == 'POST':
        action = request.form.get('action')
        student_ids = request.form.getlist('student_ids')

        if action == 'enroll':
            for sid in student_ids:
                if not Enrollment.query.filter_by(student_id=sid, course_id=course_id).first():
                    db.session.add(Enrollment(student_id=int(sid), course_id=course_id))
            db.session.commit()
            flash(f'{len(student_ids)} students enrolled.', 'success')
        elif action == 'unenroll':
            for sid in student_ids:
                e = Enrollment.query.filter_by(student_id=sid, course_id=course_id).first()
                if e:
                    db.session.delete(e)
            db.session.commit()
            flash(f'{len(student_ids)} students unenrolled.', 'success')

    enrolled_ids = [e.student_id for e in course.enrollments]
    enrolled_students = User.query.filter(User.id.in_(enrolled_ids)).all()
    all_students = User.query.filter_by(role='student').all()
    unenrolled = [s for s in all_students if s.id not in enrolled_ids]

    return render_template('manage_enrollments.html',
                           course=course, enrolled_students=enrolled_students,
                           unenrolled_students=unenrolled)


@main.route('/teacher/course/<int:course_id>/attendance', methods=['GET', 'POST'])
@login_required
@role_required('teacher')
def mark_attendance(course_id):
    course = Course.query.get_or_404(course_id)
    if course.teacher_id != current_user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('main.teacher_dashboard'))

    selected_date = request.args.get('date', date.today().isoformat())
    try:
        att_date = date.fromisoformat(selected_date)
    except ValueError:
        att_date = date.today()

    if request.method == 'POST':
        att_date_str = request.form.get('date', date.today().isoformat())
        att_date = date.fromisoformat(att_date_str)

        enrolled_ids = [e.student_id for e in course.enrollments]
        updated = 0
        for sid in enrolled_ids:
            status = request.form.get(f'status_{sid}', 'absent')
            existing = Attendance.query.filter_by(
                student_id=sid, course_id=course_id, date=att_date).first()
            if existing:
                existing.status = status
                existing.marked_at = datetime.utcnow()
            else:
                db.session.add(Attendance(
                    student_id=sid, course_id=course_id,
                    date=att_date, status=status
                ))
            updated += 1

        db.session.commit()

        # Emit real-time update
        socketio.emit('attendance_updated', {
            'course_id': course_id,
            'date': att_date_str,
            'count': updated
        })

        flash(f'Attendance saved for {updated} students on {att_date}.', 'success')
        return redirect(url_for('main.mark_attendance', course_id=course_id, date=att_date_str))

    enrolled_ids = [e.student_id for e in course.enrollments]
    students = User.query.filter(User.id.in_(enrolled_ids)).order_by(User.roll_no).all()

    existing_att = {
        a.student_id: a.status
        for a in Attendance.query.filter_by(course_id=course_id, date=att_date).all()
    }

    return render_template('mark_attendance.html',
                           course=course, students=students,
                           existing_att=existing_att, selected_date=att_date)


@main.route('/teacher/course/<int:course_id>/analytics')
@login_required
@role_required('teacher')
def course_analytics(course_id):
    course = Course.query.get_or_404(course_id)
    if course.teacher_id != current_user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('main.teacher_dashboard'))

    students = course.get_enrolled_students()
    total_classes = course.get_total_classes()

    student_analytics = []
    risk_counts = {'SAFE': 0, 'CAUTION': 0, 'WARNING': 0, 'CRITICAL': 0}

    for s in students:
        att_q = Attendance.query.filter_by(student_id=s.id, course_id=course_id)
        total = att_q.count()
        present = att_q.filter(Attendance.status.in_(['present', 'late'])).count()
        absent = total - present
        pct = s.get_attendance_percentage(course_id)
        risk, color = s.get_risk_level(course_id)
        can_miss, classes_needed = s.get_can_miss(course_id)
        risk_counts[risk] += 1

        student_analytics.append({
            'student': s, 'total': total, 'present': present,
            'absent': absent, 'pct': pct, 'risk': risk,
            'color': color, 'can_miss': can_miss, 'classes_needed': classes_needed
        })

    avg_pct = sum(x['pct'] for x in student_analytics) / len(student_analytics) if student_analytics else 0

    return render_template('course_analytics.html',
                           course=course, student_analytics=student_analytics,
                           risk_counts=risk_counts, avg_pct=avg_pct,
                           total_classes=total_classes)


@main.route('/teacher/course/<int:course_id>/export')
@login_required
@role_required('teacher')
def export_attendance(course_id):
    course = Course.query.get_or_404(course_id)
    if course.teacher_id != current_user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('main.teacher_dashboard'))

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{course.code} Attendance"

        # Header
        header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        headers = ['Roll No', 'Student Name', 'Total Classes', 'Present', 'Absent', 'Percentage', 'Risk Level']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        students = course.get_enrolled_students()
        for row, s in enumerate(students, 2):
            att_q = Attendance.query.filter_by(student_id=s.id, course_id=course_id)
            total = att_q.count()
            present = att_q.filter(Attendance.status.in_(['present', 'late'])).count()
            absent = total - present
            pct = round((present / total * 100), 1) if total > 0 else 0
            risk, _ = s.get_risk_level(course_id)

            ws.cell(row=row, column=1, value=s.roll_no or 'N/A')
            ws.cell(row=row, column=2, value=s.username)
            ws.cell(row=row, column=3, value=total)
            ws.cell(row=row, column=4, value=present)
            ws.cell(row=row, column=5, value=absent)
            ws.cell(row=row, column=6, value=f"{pct}%")
            cell = ws.cell(row=row, column=7, value=risk)

            # Color code by risk
            colors = {'SAFE': 'd4edda', 'CAUTION': 'fff3cd', 'WARNING': 'fde5d4', 'CRITICAL': 'f8d7da'}
            cell.fill = PatternFill(start_color=colors.get(risk, 'ffffff'),
                                    end_color=colors.get(risk, 'ffffff'), fill_type='solid')

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f'{course.code}_attendance.xlsx')
    except ImportError:
        flash('openpyxl not installed. Cannot export Excel.', 'danger')
        return redirect(url_for('main.course_analytics', course_id=course_id))


# ─── QR Code Routes ────────────────────────────────────────────
@main.route('/teacher/course/<int:course_id>/qr')
@login_required
@role_required('teacher')
def qr_display(course_id):
    course = Course.query.get_or_404(course_id)
    if course.teacher_id != current_user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('main.teacher_dashboard'))

    from app.utils.qr_handler import generate_qr_token, generate_qr_image
    token = generate_qr_token(course_id, current_user.id)
    qr_img = generate_qr_image(token, course.name)

    return render_template('qr_display.html', course=course, token=token,
                           qr_img=qr_img, expiry_minutes=5)


@main.route('/scan-qr', methods=['GET', 'POST'])
@login_required
@role_required('student')
def scan_qr():
    if request.method == 'POST':
        token = request.form.get('token', '').strip()
        from app.utils.qr_handler import validate_qr_token
        payload, error = validate_qr_token(token)
        if error:
            flash(error, 'danger')
            return render_template('scan_qr.html')

        course_id = payload['course_id']
        # Check if student is enrolled
        enrollment = Enrollment.query.filter_by(
            student_id=current_user.id, course_id=course_id).first()
        if not enrollment:
            flash('You are not enrolled in this course.', 'danger')
            return render_template('scan_qr.html')

        today = date.today()
        existing = Attendance.query.filter_by(
            student_id=current_user.id, course_id=course_id, date=today).first()
        if existing:
            flash('Attendance already marked for today.', 'info')
        else:
            db.session.add(Attendance(
                student_id=current_user.id, course_id=course_id,
                date=today, status='present'
            ))
            db.session.commit()
            flash('✅ Attendance marked successfully!', 'success')

        return redirect(url_for('main.student_dashboard'))

    return render_template('scan_qr.html')


@main.route('/api/qr/scan', methods=['POST'])
@login_required
def api_qr_scan():
    """API endpoint for QR scanning from camera"""
    data = request.get_json()
    token = data.get('token', '')

    from app.utils.qr_handler import validate_qr_token
    payload, error = validate_qr_token(token.replace('ATTENDANCE:', ''))
    if error:
        return jsonify({'success': False, 'message': error})

    course_id = payload['course_id']
    if current_user.role == 'student':
        enrollment = Enrollment.query.filter_by(
            student_id=current_user.id, course_id=course_id).first()
        if not enrollment:
            return jsonify({'success': False, 'message': 'Not enrolled in this course'})

        today = date.today()
        existing = Attendance.query.filter_by(
            student_id=current_user.id, course_id=course_id, date=today).first()
        if existing:
            return jsonify({'success': False, 'message': 'Attendance already marked'})

        db.session.add(Attendance(
            student_id=current_user.id, course_id=course_id,
            date=today, status='present'
        ))
        db.session.commit()

        # Notify teacher via WebSocket
        course = Course.query.get(course_id)
        socketio.emit('student_scanned', {
            'student_name': current_user.username,
            'roll_no': current_user.roll_no,
            'course_id': course_id,
            'time': datetime.now().strftime('%H:%M:%S')
        }, room=f'course_{course_id}')

        return jsonify({'success': True, 'message': 'Attendance marked!'})

    return jsonify({'success': False, 'message': 'Invalid role'})


# ─── Alert Routes ──────────────────────────────────────────────
@main.route('/teacher/alert/<int:student_id>/<int:course_id>', methods=['POST'])
@login_required
@role_required('teacher')
def send_alert(student_id, course_id):
    student = User.query.get_or_404(student_id)
    course = Course.query.get_or_404(course_id)
    pct = student.get_attendance_percentage(course_id)
    _, classes_needed = student.get_can_miss(course_id)

    results = []
    if student.parent_phone:
        from app.utils.whatsapp import send_low_attendance_alert, send_critical_alert
        if pct < 65:
            ok, msg = send_critical_alert(student, course, pct, classes_needed)
        else:
            ok, msg = send_low_attendance_alert(student, course, pct)
        results.append(f"WhatsApp: {'✅ Sent' if ok else f'❌ {msg}'}")

    if student.parent_email:
        from app.utils.email_handler import send_low_attendance_email
        ok, msg = send_low_attendance_email(student, course, pct)
        results.append(f"Email: {'✅ Sent' if ok else f'❌ {msg}'}")

    if not results:
        flash('No parent contact info found.', 'warning')
    else:
        flash(' | '.join(results), 'info')

    return redirect(url_for('main.course_analytics', course_id=course_id))


# ─── Student Routes ────────────────────────────────────────────
@main.route('/student')
@login_required
@role_required('student')
def student_dashboard():
    enrollments = Enrollment.query.filter_by(student_id=current_user.id).all()
    course_ids = [e.course_id for e in enrollments]
    courses = Course.query.filter(Course.id.in_(course_ids)).all()

    course_stats = []
    for c in courses:
        pct = current_user.get_attendance_percentage(c.id)
        risk, color = current_user.get_risk_level(c.id)
        can_miss, classes_needed = current_user.get_can_miss(c.id)
        att_q = Attendance.query.filter_by(student_id=current_user.id, course_id=c.id)
        total = att_q.count()
        present = att_q.filter(Attendance.status.in_(['present', 'late'])).count()
        course_stats.append({
            'course': c, 'pct': pct, 'risk': risk, 'color': color,
            'can_miss': can_miss, 'classes_needed': classes_needed,
            'total': total, 'present': present
        })

    overall_pct = sum(x['pct'] for x in course_stats) / len(course_stats) if course_stats else 0
    overall_risk, overall_color = current_user.get_risk_level()

    # Recent attendance (last 7 days)
    week_ago = date.today() - timedelta(days=7)
    recent_att = Attendance.query.filter(
        Attendance.student_id == current_user.id,
        Attendance.date >= week_ago
    ).order_by(Attendance.date.desc()).all()

    return render_template('student_dashboard.html',
                           course_stats=course_stats, overall_pct=overall_pct,
                           overall_risk=overall_risk, overall_color=overall_color,
                           recent_att=recent_att)


# ─── WebSocket Events ──────────────────────────────────────────
@socketio.on('join_course')
def on_join_course(data):
    from flask_socketio import join_room
    room = f"course_{data['course_id']}"
    join_room(room)


@socketio.on('leave_course')
def on_leave_course(data):
    from flask_socketio import leave_room
    room = f"course_{data['course_id']}"
    leave_room(room)


# ─── API Endpoints ─────────────────────────────────────────────
@main.route('/api/analytics/overview')
@login_required
@role_required('admin')
def api_analytics_overview():
    """Return system-wide analytics as JSON for charts"""
    students = User.query.filter_by(role='student').all()
    risk_counts = {'SAFE': 0, 'CAUTION': 0, 'WARNING': 0, 'CRITICAL': 0}
    pct_distribution = []

    for s in students:
        pct = s.get_attendance_percentage()
        risk, _ = s.get_risk_level()
        risk_counts[risk] += 1
        pct_distribution.append(pct)

    return jsonify({
        'risk_counts': risk_counts,
        'avg_attendance': sum(pct_distribution) / len(pct_distribution) if pct_distribution else 0,
        'total_students': len(students)
    })


@main.route('/teacher/course/<int:course_id>/student/<int:student_id>/prediction')
@login_required
@role_required('teacher')
def student_prediction(course_id, student_id):
    """Show ML attendance prediction for a specific student."""
    course = Course.query.get_or_404(course_id)
    if course.teacher_id != current_user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('main.teacher_dashboard'))

    student = User.query.get_or_404(student_id)

    from app.utils.analytics import get_attendance_trend, predict_future_attendance, get_weekly_pattern
    trend = get_attendance_trend(student_id, course_id, db, Attendance)
    prediction = predict_future_attendance(student_id, course_id, db, Attendance)
    weekly = get_weekly_pattern(student_id, course_id, db, Attendance)

    return render_template('student_prediction.html',
                           course=course, student=student,
                           trend=json.dumps(trend),
                           prediction=prediction,
                           weekly=json.dumps(weekly))


@main.route('/api/student/<int:student_id>/trend/<int:course_id>')
@login_required
def api_student_trend(student_id, course_id):
    """Return attendance trend data as JSON for Chart.js."""
    # Students can only see their own data
    if current_user.role == 'student' and current_user.id != student_id:
        return jsonify({'error': 'Access denied'}), 403

    from app.utils.analytics import get_attendance_trend
    trend = get_attendance_trend(student_id, course_id, db, Attendance)
    return jsonify(trend)


@main.route('/teacher/course/<int:course_id>/live')
@login_required
@role_required('teacher')
def live_dashboard(course_id):
    """Real-time live attendance dashboard via WebSockets."""
    course = Course.query.get_or_404(course_id)
    if course.teacher_id != current_user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('main.teacher_dashboard'))

    enrolled_ids = [e.student_id for e in course.enrollments]
    students = User.query.filter(User.id.in_(enrolled_ids)).order_by(User.roll_no).all()
    return render_template('live_dashboard.html', course=course, students=students)


# ═══════════════════════════════════════════════════════════════
# FACE RECOGNITION ROUTES
# ═══════════════════════════════════════════════════════════════

@main.route('/face/register')
@login_required
def face_register():
    """Face registration page — for students and teachers."""
    from app.utils.face_recognition_handler import is_face_registered

    already_registered = is_face_registered(current_user.id)

    # For teachers: show all students so they can register on behalf
    students = []
    if current_user.role in ('teacher', 'admin'):
        all_students = User.query.filter_by(role='student').all()
        students = [{
            'id': s.id,
            'username': s.username,
            'roll_no': s.roll_no,
            'face_registered': is_face_registered(s.id)
        } for s in all_students]

    return render_template('face_register.html',
                           already_registered=already_registered,
                           students=students)


@main.route('/teacher/course/<int:course_id>/face-attendance')
@login_required
@role_required('teacher')
def face_attendance(course_id):
    """Face recognition attendance page for teachers."""
    from app.utils.face_recognition_handler import is_face_registered, get_registered_count

    course = Course.query.get_or_404(course_id)
    if course.teacher_id != current_user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('main.teacher_dashboard'))

    enrolled_ids = [e.student_id for e in course.enrollments]
    students = User.query.filter(User.id.in_(enrolled_ids)).order_by(User.roll_no).all()

    students_data = [{
        'id': s.id,
        'name': s.username,
        'roll_no': s.roll_no or 'N/A',
        'face_registered': is_face_registered(s.id)
    } for s in students]

    registered_count = get_registered_count(enrolled_ids)

    return render_template('face_attendance.html',
                           course=course,
                           students=students,
                           students_json=json.dumps(students_data),
                           total_enrolled=len(students),
                           registered_count=registered_count,
                           today=date.today().isoformat())


# ─── Face API Endpoints ────────────────────────────────────────

@main.route('/api/face/register', methods=['POST'])
@login_required
def api_face_register():
    """Save face images for a user."""
    from app.utils.face_recognition_handler import save_face_image, delete_face_data
    import os

    data = request.get_json()
    student_id = data.get('student_id', current_user.id)
    images = data.get('images', [])

    # Security: students can only register their own face
    if current_user.role == 'student' and student_id != current_user.id:
        return jsonify({'success': False, 'message': 'Access denied'})

    if not images:
        return jsonify({'success': False, 'message': 'No images provided'})

    # Delete old face data first (re-registration)
    delete_face_data(student_id)

    saved = 0
    errors = []
    for i, img_data in enumerate(images[:5]):  # max 5 photos
        ok, result = save_face_image(student_id, img_data, filename=f'face_{i+1}.jpg')
        if ok:
            saved += 1
        else:
            errors.append(result)

    if saved == 0:
        return jsonify({
            'success': False,
            'message': errors[0] if errors else 'No valid face images saved'
        })

    return jsonify({
        'success': True,
        'saved': saved,
        'message': f'{saved} face images registered successfully'
    })


@main.route('/api/face/recognize', methods=['POST'])
@login_required
@role_required('teacher')
def api_face_recognize():
    """Recognize faces in a classroom image."""
    from app.utils.face_recognition_handler import recognize_faces_in_image

    data = request.get_json()
    image_data = data.get('image')
    course_id = data.get('course_id')

    if not image_data or not course_id:
        return jsonify({'error': 'Missing image or course_id'})

    course = Course.query.get_or_404(course_id)
    if course.teacher_id != current_user.id:
        return jsonify({'error': 'Access denied'})

    enrolled_ids = [e.student_id for e in course.enrollments]

    result = recognize_faces_in_image(image_data, enrolled_ids, model_name='VGG-Face')
    return jsonify(result)


@main.route('/api/face/save-attendance', methods=['POST'])
@login_required
@role_required('teacher')
def api_face_save_attendance():
    """Save attendance for recognized students."""
    data = request.get_json()
    course_id = data.get('course_id')
    att_date_str = data.get('date', date.today().isoformat())
    recognized_ids = data.get('recognized_ids', [])

    course = Course.query.get_or_404(course_id)
    if course.teacher_id != current_user.id:
        return jsonify({'success': False, 'message': 'Access denied'})

    try:
        att_date = date.fromisoformat(att_date_str)
    except ValueError:
        att_date = date.today()

    enrolled_ids = [e.student_id for e in course.enrollments]
    marked = 0

    for sid in enrolled_ids:
        status = 'present' if sid in recognized_ids else 'absent'
        existing = Attendance.query.filter_by(
            student_id=sid, course_id=course_id, date=att_date).first()
        if existing:
            existing.status = status
            existing.marked_at = datetime.utcnow()
        else:
            db.session.add(Attendance(
                student_id=sid, course_id=course_id,
                date=att_date, status=status
            ))
        if status == 'present':
            marked += 1

    db.session.commit()

    socketio.emit('attendance_updated', {
        'course_id': course_id,
        'date': att_date_str,
        'count': len(enrolled_ids),
        'method': 'face_recognition'
    })

    return jsonify({'success': True, 'marked': marked, 'total': len(enrolled_ids)})


@main.route('/api/face/delete', methods=['POST'])
@login_required
def api_face_delete():
    """Delete face data for current user."""
    from app.utils.face_recognition_handler import delete_face_data
    delete_face_data(current_user.id)
    flash('Face data deleted successfully.', 'success')
    return redirect(url_for('main.face_register'))


@main.route('/face/delete', methods=['POST'])
@login_required
def delete_face_data():
    """Delete face data route (form submit)."""
    from app.utils.face_recognition_handler import delete_face_data as _delete
    _delete(current_user.id)
    flash('Face data deleted.', 'success')
    return redirect(url_for('main.face_register'))



# ═══════════════════════════════════════════════════════════════
# AI CHATBOT ROUTES
# ═══════════════════════════════════════════════════════════════

@main.route('/chatbot')
@login_required
def chatbot():
    """AI chatbot page."""
    from app.utils.chatbot import QUICK_REPLIES
    quick_replies = QUICK_REPLIES.get(current_user.role, QUICK_REPLIES['student'])
    return render_template('chatbot.html', quick_replies=quick_replies)


@main.route('/api/chatbot', methods=['POST'])
@login_required
def api_chatbot():
    """Handle chatbot message and return AI response."""
    from app.utils.chatbot import build_system_prompt, get_chatbot_response

    data = request.get_json()
    user_message = data.get('message', '').strip()
    history = data.get('history', [])

    if not user_message:
        return jsonify({'error': 'Empty message'})

    if len(user_message) > 500:
        return jsonify({'error': 'Message too long (max 500 characters)'})

    # Build live system prompt with DB context
    system_prompt = build_system_prompt(
        current_user, db, User, Course, Enrollment, Attendance
    )

    # Get AI response
    response, error = get_chatbot_response(user_message, history, system_prompt)

    if error:
        return jsonify({'error': error})

    return jsonify({'response': response})



# ═══════════════════════════════════════════════════════════════
# PARENT PORTAL ROUTES
# ═══════════════════════════════════════════════════════════════

@main.route('/parent/register', methods=['GET', 'POST'])
def parent_register():
    """Parent self-registration page."""
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        phone = request.form.get('phone', '').strip()
        roll_no = request.form.get('roll_no', '').strip()

        if User.query.filter_by(username=username).first():
            flash('Username already exists.', 'danger')
            return render_template('parent_register.html')
        if User.query.filter_by(email=email).first():
            flash('Email already registered.', 'danger')
            return render_template('parent_register.html')

        # Find child by roll number
        child = User.query.filter_by(roll_no=roll_no, role='student').first()
        if not child:
            flash(f'No student found with roll number "{roll_no}". Please check and try again.', 'danger')
            return render_template('parent_register.html')

        # Create parent account
        parent = User(username=username, email=email, role='parent',
                      parent_phone=phone, is_active=True)
        parent.set_password(password)
        db.session.add(parent)
        db.session.flush()

        # Link to child
        from app.models import ParentStudent
        link = ParentStudent(parent_id=parent.id, student_id=child.id, verified=True)
        db.session.add(link)
        db.session.commit()

        flash(f'Account created! You are now linked to {child.username}.', 'success')
        login_user(parent)
        return redirect(url_for('main.parent_dashboard'))

    return render_template('parent_register.html')


@main.route('/parent')
@login_required
@role_required('parent')
def parent_dashboard():
    """Main parent portal dashboard."""
    from app.models import ParentStudent, TeacherMessage
    from datetime import timedelta

    children = current_user.get_linked_children()
    children_data = []

    for child in children:
        enrollments = Enrollment.query.filter_by(student_id=child.id).all()
        course_stats = []
        total_present = 0

        for e in enrollments:
            course = Course.query.get(e.course_id)
            if not course:
                continue
            att_q = Attendance.query.filter_by(student_id=child.id, course_id=course.id)
            total = att_q.count()
            present = att_q.filter(Attendance.status.in_(['present', 'late'])).count()
            absent = total - present
            pct = round((present / total * 100), 1) if total > 0 else 0
            risk, _ = child.get_risk_level(course.id)
            can_miss, classes_needed = child.get_can_miss(course.id)
            total_present += present
            course_stats.append({
                'course': course, 'total': total, 'present': present,
                'absent': absent, 'pct': pct, 'risk': risk,
                'can_miss': can_miss, 'classes_needed': classes_needed
            })

        overall_pct = round(sum(c['pct'] for c in course_stats) / len(course_stats), 1) if course_stats else 0
        overall_risk, _ = child.get_risk_level()

        # Build 30-day calendar
        today = date.today()
        calendar_days = []
        # Find offset to start on Monday
        start = today - timedelta(days=29)
        weekday_offset = start.weekday()  # 0=Mon
        for _ in range(weekday_offset):
            calendar_days.append({'day': '', 'css': 'empty', 'label': ''})

        for i in range(30):
            d = start + timedelta(days=i)
            if d.weekday() >= 5:
                calendar_days.append({'day': d.day, 'css': 'weekend', 'label': d.strftime('%a %d %b')})
                continue
            # Check attendance across all courses for this day
            att_records = Attendance.query.filter_by(student_id=child.id, date=d).all()
            if not att_records:
                calendar_days.append({'day': d.day, 'css': 'empty', 'label': d.strftime('%d %b — no class')})
            else:
                statuses = [r.status for r in att_records]
                if 'present' in statuses or 'late' in statuses:
                    calendar_days.append({'day': d.day, 'css': 'present', 'label': d.strftime('%d %b — Present')})
                else:
                    calendar_days.append({'day': d.day, 'css': 'absent', 'label': d.strftime('%d %b — Absent')})

        # Get recent messages
        messages = TeacherMessage.query.filter_by(student_id=child.id).order_by(
            TeacherMessage.created_at.asc()).limit(20).all()

        children_data.append({
            'child': child,
            'courses': course_stats,
            'overall_pct': overall_pct,
            'overall_risk': overall_risk,
            'total_present': total_present,
            'calendar': calendar_days,
            'messages': messages
        })

    return render_template('parent_dashboard.html',
                           children=children, children_data=children_data)


@main.route('/parent/link', methods=['GET', 'POST'])
@login_required
@role_required('parent')
def parent_link_student():
    """Link additional child to parent account."""
    from app.models import ParentStudent

    if request.method == 'POST':
        roll_no = request.form.get('roll_no', '').strip()
        child = User.query.filter_by(roll_no=roll_no, role='student').first()
        if not child:
            flash(f'No student found with roll number "{roll_no}".', 'danger')
        else:
            existing = ParentStudent.query.filter_by(
                parent_id=current_user.id, student_id=child.id).first()
            if existing:
                flash(f'{child.username} is already linked to your account.', 'info')
            else:
                db.session.add(ParentStudent(
                    parent_id=current_user.id, student_id=child.id, verified=True))
                db.session.commit()
                flash(f'Successfully linked to {child.username}!', 'success')
        return redirect(url_for('main.parent_link_student'))

    linked_children = current_user.get_linked_children()
    return render_template('parent_link.html', linked_children=linked_children)


@main.route('/parent/message/<int:teacher_id>/<int:student_id>', methods=['GET'])
@main.route('/parent/message/<int:teacher_id>/<int:student_id>/<int:course_id>', methods=['GET'])
@login_required
@role_required('parent')
def parent_message_teacher(teacher_id, student_id, course_id=None):
    """View message thread between parent and teacher."""
    from app.models import TeacherMessage

    teacher = User.query.get_or_404(teacher_id)
    student = User.query.get_or_404(student_id)
    course = Course.query.get(course_id) if course_id else None

    # Verify this parent is linked to this student
    from app.models import ParentStudent
    link = ParentStudent.query.filter_by(
        parent_id=current_user.id, student_id=student_id).first()
    if not link:
        flash('Access denied.', 'danger')
        return redirect(url_for('main.parent_dashboard'))

    messages = TeacherMessage.query.filter(
        TeacherMessage.student_id == student_id,
        db.or_(
            db.and_(TeacherMessage.sender_id == current_user.id,
                    TeacherMessage.receiver_id == teacher_id),
            db.and_(TeacherMessage.sender_id == teacher_id,
                    TeacherMessage.receiver_id == current_user.id)
        )
    ).order_by(TeacherMessage.created_at.asc()).all()

    # Mark messages as read
    for msg in messages:
        if msg.receiver_id == current_user.id and not msg.is_read:
            msg.is_read = True
    db.session.commit()

    return render_template('parent_message.html',
                           teacher=teacher, student=student,
                           course=course, messages=messages)


@main.route('/parent/message/send', methods=['POST'])
@login_required
@role_required('parent')
def parent_send_message():
    """Send a message to a teacher."""
    from app.models import TeacherMessage, ParentStudent

    teacher_id = int(request.form.get('teacher_id'))
    student_id = int(request.form.get('student_id'))
    course_id = request.form.get('course_id')
    message_text = request.form.get('message', '').strip()

    if not message_text:
        flash('Message cannot be empty.', 'danger')
        return redirect(url_for('main.parent_dashboard'))

    # Verify link
    link = ParentStudent.query.filter_by(
        parent_id=current_user.id, student_id=student_id).first()
    if not link:
        flash('Access denied.', 'danger')
        return redirect(url_for('main.parent_dashboard'))

    msg = TeacherMessage(
        sender_id=current_user.id,
        receiver_id=teacher_id,
        student_id=student_id,
        course_id=int(course_id) if course_id else None,
        message=message_text
    )
    db.session.add(msg)
    db.session.commit()
    flash('Message sent to teacher.', 'success')

    cid = int(course_id) if course_id else None
    return redirect(url_for('main.parent_message_teacher',
                            teacher_id=teacher_id, student_id=student_id,
                            course_id=cid))


@main.route('/parent/report/<int:student_id>')
@login_required
@role_required('parent')
def parent_download_report(student_id):
    """Download Excel attendance report for a child."""
    from app.models import ParentStudent

    link = ParentStudent.query.filter_by(
        parent_id=current_user.id, student_id=student_id).first()
    if not link:
        flash('Access denied.', 'danger')
        return redirect(url_for('main.parent_dashboard'))

    student = User.query.get_or_404(student_id)

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Attendance Report — {student.username} ({student.roll_no or 'N/A'})"
        ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:F2')
        ws['A2'] = f"Generated on {date.today().strftime('%d %B %Y')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A2'].font = Font(italic=True, color='666666')

        # Headers
        headers = ['Course', 'Code', 'Total Classes', 'Present', 'Absent', 'Percentage']
        hfill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.fill = hfill
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')

        enrollments = Enrollment.query.filter_by(student_id=student.id).all()
        for row, e in enumerate(enrollments, 5):
            course = Course.query.get(e.course_id)
            if not course:
                continue
            att_q = Attendance.query.filter_by(student_id=student.id, course_id=course.id)
            total = att_q.count()
            present = att_q.filter(Attendance.status.in_(['present', 'late'])).count()
            absent = total - present
            pct = round((present / total * 100), 1) if total > 0 else 0

            ws.cell(row=row, column=1, value=course.name)
            ws.cell(row=row, column=2, value=course.code)
            ws.cell(row=row, column=3, value=total)
            ws.cell(row=row, column=4, value=present)
            ws.cell(row=row, column=5, value=absent)
            pct_cell = ws.cell(row=row, column=6, value=f"{pct}%")

            color = 'd4edda' if pct >= 80 else 'fff3cd' if pct >= 75 else 'f8d7da'
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=color, end_color=color, fill_type='solid')

        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=f'{student.username}_attendance_{date.today()}.xlsx')
    except ImportError:
        flash('openpyxl not installed.', 'danger')
        return redirect(url_for('main.parent_dashboard'))


# ── Teacher: view messages from parents ───────────────────────
@main.route('/teacher/messages')
@login_required
@role_required('teacher')
def teacher_messages():
    """Teacher view of all parent messages."""
    from app.models import TeacherMessage

    messages = TeacherMessage.query.filter_by(
        receiver_id=current_user.id
    ).order_by(TeacherMessage.created_at.desc()).all()

    # Mark as read
    for msg in messages:
        if not msg.is_read:
            msg.is_read = True
    db.session.commit()

    unread_by_parent = {}
    for msg in messages:
        key = (msg.sender_id, msg.student_id)
        if key not in unread_by_parent:
            unread_by_parent[key] = msg

    return render_template('teacher_messages.html',
                           messages=messages,
                           grouped=list(unread_by_parent.values()))


@main.route('/teacher/messages/reply', methods=['POST'])
@login_required
@role_required('teacher')
def teacher_reply_message():
    """Teacher replies to a parent message."""
    from app.models import TeacherMessage

    parent_id = int(request.form.get('parent_id'))
    student_id = int(request.form.get('student_id'))
    course_id = request.form.get('course_id')
    message_text = request.form.get('message', '').strip()

    if message_text:
        msg = TeacherMessage(
            sender_id=current_user.id,
            receiver_id=parent_id,
            student_id=student_id,
            course_id=int(course_id) if course_id else None,
            message=message_text
        )
        db.session.add(msg)
        db.session.commit()
        flash('Reply sent.', 'success')

    return redirect(url_for('main.teacher_messages'))

